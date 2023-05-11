# Parse Fasta
from bs4 import BeautifulSoup
import requests

# Creating Excel
import openpyxl
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

# Time
import time

from numba import njit, prange


class Checker:
    def __init__(self, protein_name):
        self.protein_name = protein_name

    def url_existence(self):
        url = f'https://rest.uniprot.org/uniprotkb/{self.protein_name}.fasta'
        session = requests.Session()
        html_text = session.get(url).text
        soup = BeautifulSoup(html_text, 'lxml')
        if soup.p.string[0: 14] != 'Error messages':
            return True
        return False


class Parser:
    def __init__(self, protein_name):
        self.protein_name = protein_name

    def parsing_uniprot(self):
        data = ["None", "None", "None", "None"]
        try:
            urluniprot = f'https://rest.uniprot.org/uniprotkb/{self.protein_name}.json'
            session = requests.Session()
            response = session.get(urluniprot)
            data = response.json()
            return data

        except BaseException:
            return data

        finally:
            return data

    def parsing_fasta(self):
        count_try = 0
        data = ''
        while count_try != 3:
            try:
                url = f'https://rest.uniprot.org/uniprotkb/{self.protein_name}.fasta'
                session = requests.Session()
                html_text = session.get(url).text
                soup = BeautifulSoup(html_text, 'lxml')
                data = str(soup.p.string)  # FastaFile
                return data
            except BaseException:
                count_try += 1
        return data


class InformationHandler:
    def __init__(self, uniprot_data, fasta_data):
        self.uniprot_data = uniprot_data
        self.fasta_data = fasta_data

    def fasta_processing(self):
        result_list = ['None', 'None', 'None', 'None', 'None', 'None', 'None']
        try:
            file = self.fasta_data
            position1 = file.find('|', 1)
            name_protein = file[file.find(
                '|', 1) + 1: file.find('|', position1 + 1)]
            entry_name = file[file.find(
                '|', position1 + 1) + 1: file.find(' ')]
            protein = file[file.find(' ') + 1: file.find('OS')]
            os = file[file.find('=') + 1: file.find('OX')]
            gen = file[file.find('GN') + 3: file.find('PE')]
            startchain = file.find('\n') + 1
            chain_with_space = file[startchain:len(file)]

            @njit(fastmath=True)
            def get_chain():
                chain = ''
                for i in prange(len(chain_with_space)):
                    if chain_with_space[i] != '\n':
                        chain += chain_with_space[i]
                return chain

            chain = get_chain()
            length_chain = len(chain)

            result_list = [
                name_protein,
                entry_name,
                protein,
                os,
                gen,
                length_chain,
                chain
            ]
            return result_list

        except BaseException:
            print("returned NONE")
            return result_list

    def uniprot_processing(self):
        result_list = ["None", "None", "None", "None"]
        try:
            data = self.uniprot_data
            mass_da = str(data["sequence"]["molWeight"])[0:2] + ',' + str(data["sequence"]["molWeight"])[2:len(
                str(data["sequence"]["molWeight"]))]
            status = data["entryType"]
            status = status[status.find('(') + 1: len(status) - 1]
            organism = data["organism"]["commonName"]
            existence = data["proteinExistence"][3:len(data["proteinExistence"])]
            result_list = [mass_da, status, organism, existence]
            """result_list.append(mass_da)
            result_list.append(status)
            result_list.append(organism)
            result_list.append(existence)"""

            return result_list

        except BaseException:
            print('returned None')
            return result_list


class Creater:
    def __init__(self, filename):
        self.filename = filename

    def creating_excel(self):
        config = open("cfg/User_config.txt", "r", encoding="utf-8")
        line = config.readline()
        config.close()
        save_path = line[(line.find("Save_Path:") + len("Save_Path:")): line.find("@") - 1]
        filepath = f'{save_path}/{self.filename}.xlsx'
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']
        sheet.title = 'Result'
        wb.save(filepath)


class Writer:
    def __init__(self, uniprot_result, fasta_result):
        self.uniprot_result = uniprot_result
        self.fasta_result = fasta_result

        self.columns = ['Entry identifier', 'Entry name', 'Status', 'Protein name',
                        'Organism (scientific name)', 'Organism (common name)',
                        'Gene name', 'Protein existence', 'Length', 'Mass (Da)']

        self.columns_peptides = ['Category', 'Peptide ID', 'Sequence', 'Length',
                                 'Occurrence', 'Relative (per 1000 amino acids)']

        self.columns_aminos = ['Position', 'Amino acid from the N-terminus', 'Amino acid from the C-terminus']

        config = open("cfg/User_config.txt", "r", encoding="utf-8")
        self.line = config.readline()
        self.save_path = self.line[self.line.find("Save_Path:") + len("Save_Path:"): self.line.find("@") - 1]
        config.close()

    def filling_main_info(self):
        main_data_list = [
            self.fasta_result[0],
            self.fasta_result[1],
            self.uniprot_result[1],
            self.fasta_result[2],
            self.fasta_result[3],
            self.uniprot_result[2],
            self.fasta_result[4],
            self.uniprot_result[3],
            self.fasta_result[5],
            self.uniprot_result[0]
        ]

        cfg = self.line
        cfg = cfg[cfg.find("Excel filters:") + 14: cfg.find("Excel filters:") + 14 + 19]
        for i in prange(0, 10):
            if cfg[i] == '0':
                self.columns[i] = ' '
                main_data_list[i] = ' '

        wb = load_workbook(f'{self.save_path}/{main_data_list[0]}.xlsx')
        result = wb.active

        is_checked = 0
        for i in prange(len(self.columns)):
            if self.columns[i] == ' ':
                is_checked += 1
            else:
                result.cell(row=2, column=i + 2 - is_checked).value = self.columns[i]
                result.cell(row=2, column=i + 2 - is_checked).font = Font(bold=True)

                result.cell(row=2, column=i + 2 - is_checked).fill = PatternFill(
                    start_color="FDE9D9",
                    end_color="FDE9D9",
                    fill_type="solid")

                result.cell(row=2, column=i + 2 - is_checked).alignment = Alignment(
                    horizontal='left',
                    vertical='top')

        result.cell(row=2, column=1).fill = PatternFill(
            start_color="FDE9D9",
            end_color="FDE9D9",
            fill_type="solid")

        is_checked = 0
        for i in prange(len(main_data_list)):
            if self.columns[i] == ' ':
                is_checked += 1
            else:
                result.cell(row=3, column=i + 2 - is_checked).value = main_data_list[i]
                result.cell(row=3, column=i + 2 - is_checked).font = Font(color="C00000")
                result.cell(row=3, column=i + 2 - is_checked).alignment = Alignment(horizontal='left')

        for col in result.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except BaseException:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            result.column_dimensions[column].width = adjusted_width

        wb.save(f"{self.save_path}/{main_data_list[0]}.xlsx")

    def filling_peptides_info(self):
        cfg = self.line
        cfg = cfg[cfg.find("Excel filters:") + 14: cfg.find("Excel filters:") + 14 + 19]
        for i in prange(10, 16):
            if cfg[i] == '0':
                self.columns_peptides[i - 10] = ' '

        for i in prange(16, len(cfg)):
            if cfg[i] == '0':
                self.columns_aminos[i - 16] = ' '

        chain = self.fasta_result[6]
        wb = load_workbook(f'{self.save_path}/{self.fasta_result[0]}.xlsx')
        result = wb.active

        file_peptides = open('cfg/Peptides.txt', "r")
        spisok = file_peptides.readline()
        peptides_array = list(spisok.split())
        r = 5
        pep_flag = False
        for i in prange(len(peptides_array)):
            count_peptide = 0
            if peptides_array[i] in chain:
                pep_flag = True
                first_r = r
                not_checked = 0
                for col in prange(len(self.columns_peptides)):
                    if self.columns_peptides[col] == ' ':
                        not_checked += 1
                    else:
                        result.cell(row=r, column=col + 2 - not_checked).value = self.columns_peptides[col]
                        result.cell(row=r, column=col + 2 - not_checked).font = Font(bold=True)
                        result.cell(row=r, column=col + 2 - not_checked).fill = PatternFill(
                            start_color="DAEEF3",
                            end_color="DAEEF3",
                            fill_type="solid")
                        result.cell(row=r, column=col + 2 - not_checked).alignment = Alignment(
                            horizontal='left',
                            vertical='top')
                        result.row_dimensions[r].height = 30
                sdvig = not_checked
                not_checked = 0
                position_col = []
                for col in prange(len(self.columns_aminos)):
                    if self.columns_aminos[col] == ' ':
                        not_checked += 1
                        position_col.append(-1)
                    else:
                        result.cell(row=r, column=col + 8 - sdvig - not_checked).value = self.columns_aminos[col]
                        result.cell(row=r, column=col + 8 - sdvig - not_checked).font = Font(bold=True)
                        result.cell(row=r, column=col + 8 - sdvig - not_checked).fill = PatternFill(
                            start_color="EBF1DE",
                            end_color="EBF1DE",
                            fill_type="solid")
                        result.cell(row=r, column=col + 8 - sdvig - not_checked).alignment = Alignment(
                            horizontal='left',
                            vertical='top')
                        result.row_dimensions[r].height = 30
                        position_col.append(col + 8 - sdvig - not_checked)

                result.cell(row=r, column=1).fill = PatternFill(
                    start_color="DAEEF3",
                    end_color="DAEEF3",
                    fill_type="solid")

                for j in prange(len(chain)):
                    if chain[j: j + len(peptides_array[i])
                       ] == peptides_array[i]:
                        count_peptide += 1
                        leftpos = j + 1
                        rigthpos = j + len(peptides_array[i])
                        if j == 0:
                            nter = 'None'
                            cter = chain[j + len(peptides_array[i])]
                        elif j == len(chain) - 1:
                            nter = chain[j - 1]
                            cter = 'None'
                        else:
                            nter = chain[j - 1]
                            cter = chain[j + len(peptides_array[i])]
                        if position_col[0] != -1:
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[0]).value = f'{leftpos}-{rigthpos}'
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[0]).alignment = Alignment(
                                horizontal='left',
                                vertical='top')
                        if position_col[1] != -1:
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[1]).value = nter
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[1]).alignment = Alignment(
                                horizontal='left',
                                vertical='top')
                        if position_col[2] != -1:
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[2]).value = cter
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[2]).alignment = Alignment(
                                horizontal='left',
                                vertical='top')
                        j += len(peptides_array[i])

                category = 1200
                peptide_id = i + 1
                relative = round((count_peptide * 1000) / len(chain), 2)
                peptide = peptides_array[i]
                length_peptide = len(peptide)
                value_list = [category, peptide_id, peptide, length_peptide, count_peptide, relative]

                cfg = self.line
                cfg = cfg[cfg.find("Excel filters:") + 14: cfg.find("Excel filters:") + 14 + 19]

                for ind in prange(10, 16):
                    if cfg[ind] == '0':
                        value_list[ind - 10] = ' '

                is_checked = 0
                for val in prange(len(value_list)):
                    if value_list[val] == ' ':
                        is_checked += 1
                    else:
                        result.cell(row=first_r + 1, column=val + 2 - is_checked).value = value_list[val]
                        result.cell(row=first_r + 1, column=val + 2 - is_checked).alignment = Alignment(
                            horizontal='left',
                            vertical='top'
                        )

                r += count_peptide + 1

        if not pep_flag:
            result.cell(row=r + 2, column=2).value = 'Not peptides in file'
            result.cell(row=r + 2, column=2).font = Font(bold=True)

        for col in result.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except BaseException:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            result.column_dimensions[column].width = adjusted_width

        result.column_dimensions['A'].width = 4
        wb.save(f"{self.save_path}/{self.fasta_result[0]}.xlsx")


def main():
    file_prot = open("cfg/Proteins.txt", "r")
    #Settings.updateValue(100)
    proteins = file_prot.readline()
    file_prot.close()
    names = list(proteins.split(' '))
    #Ui_MainWindow.updateProgress(50)
    count = len(names)
    print(names)
    num_of_pep = 0
    errors_cnt = 0
    for protein in names:
        p = int((100 // count) * num_of_pep)
        num_of_pep += 1
        print(f"{p + (10 // count) * num_of_pep}%")
        protein.upper()
        check = Checker(protein_name=protein)
        print(f"{p + (15 // count) * num_of_pep}%")
        if (protein != ' ' or '') and check.url_existence():
            print(f"{p + (30 // count) * num_of_pep}%")
            parser = Parser(protein_name=protein)
            fasta = parser.parsing_fasta()
            print(f"{p + (40 // count) * num_of_pep}%")
            # print("Fasta parse - OK")
            uniprot = parser.parsing_uniprot()
            # print("Uniprot parse - OK")
            print(f"{p + (50 // count) * num_of_pep}%")
            handler = InformationHandler(uniprot_data=uniprot, fasta_data=fasta)
            fasta_result = handler.fasta_processing()
            uniprot_result = handler.uniprot_processing()
            print(f"{p + (60 // count) * num_of_pep}%")
            creater = Creater(protein)
            creater.creating_excel()
            # print("Create EXCEL - OK")
            print(f"{p + (80 // count) * num_of_pep}%")
            writer = Writer(uniprot_result=uniprot_result, fasta_result=fasta_result)
            writer.filling_main_info()
            # print("main info - OK")
            print(f"{p + (90 // count) * num_of_pep}%")
            writer.filling_peptides_info()
            # print("pep info - OK")
            # print(f'{peptide} ready')
            print(f"{int((100 // count) * num_of_pep)}%")
        else:
            errors_cnt += 1
            if errors_cnt == 1:
                logs = open("cfg/Log error.txt", "w")
                logs.write(f"{protein} ")
            else:
                logs = open("cfg/Log error.txt", "a")
                logs.write(f"{protein} ")

try:
    main()
except:
    pass
exit()




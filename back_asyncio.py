# Parse Fasta
#import requests

# Creating Excel
import openpyxl
from openpyxl.styles import Font
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Alignment

from numba import prange

import asyncio
import aiohttp
import threading

import json
from typing import List


class Creater:
    def __init__(self, filename):
        self.filename = filename

    def creating_excel(self):
        with open("config.json", "r") as cfg:
            config_data = json.load(cfg)
            save_path = config_data["savePath"]["value"]
        filepath = f'{save_path}/{self.filename}.xlsx'
        wb = openpyxl.Workbook()
        sheet = wb['Sheet']
        sheet.title = 'Result'
        wb.save(filepath)


class Writer:
    def __init__(self, uniprot_result: dict, sequence: str):
        self.columns = ['Entry identifier', 'Entry name', 'Status', 'Protein name',
                        'Organism (scientific name)', 'Organism (common name)',
                        'Gene name', 'Protein existence', 'Length', 'Mass (Da)']

        self.columns_peptides = ['Category', 'Peptide ID', 'Sequence', 'Length',
                                 'Occurrence', 'Relative (per 1000 amino acids)']

        self.columns_aminos = ['Position', 'Amino acid from the N-terminus', 'Amino acid from the C-terminus']

        self.uniprot_result = uniprot_result
        self.sequence = sequence

        with open("config.json", "r") as cfg:
            config_data = json.load(cfg)

        self.save_path = config_data["savePath"]["value"]
        self.excelFilters = config_data["excelFilters"]
        self.proteins = config_data["proteins"]["value"]
        self.peptides = config_data["peptides"]["value"]

    def filling_main_info(self):
        for i, key in enumerate(self.excelFilters.keys(), start=0):
            if i == 10:
                break
            if not(self.excelFilters[key]):
                self.columns[i] = ' '

        wb = load_workbook(f'{self.save_path}/{self.uniprot_result["proteinName"]}.xlsx')
        result = wb.active

        is_checked = 0
        for i in prange(len(self.columns)):
            if self.columns[i] == ' ':
                is_checked += 1
            else:
                result.cell(row=2, column=i + 2 - is_checked).value = self.columns[i]
                result.cell(row=2, column=i + 2 - is_checked).font = Font(bold=True)

                result.cell(row=2, column=i + 2 - is_checked).fill = PatternFill(
                    start_color="FDE9D9",
                    end_color="FDE9D9",
                    fill_type="solid")

                result.cell(row=2, column=i + 2 - is_checked).alignment = Alignment(
                    horizontal='left',
                    vertical='top')

        result.cell(row=2, column=1).fill = PatternFill(
            start_color="FDE9D9",
            end_color="FDE9D9",
            fill_type="solid")

        is_checked = 0
        for i, key in enumerate(self.uniprot_result.keys(), start=0):
            if self.columns[i] == ' ':
                is_checked += 1
            else:
                result.cell(row=3, column=i + 2 - is_checked).value = self.uniprot_result[key]
                result.cell(row=3, column=i + 2 - is_checked).font = Font(color="C00000")
                result.cell(row=3, column=i + 2 - is_checked).alignment = Alignment(horizontal='left')

        for col in result.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except BaseException:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            result.column_dimensions[column].width = adjusted_width

        wb.save(f'{self.save_path}/{self.uniprot_result["proteinName"]}.xlsx')

    def filling_peptides_info(self):
        for i, key in enumerate(self.excelFilters.keys(), start=0):
            if (i > 9 and i < 16) and not(self.excelFilters[key]):
                self.columns_peptides[i - 10] = ' '
            elif i >= 16 and not(self.excelFilters[key]):
                self.columns_aminos[i - 16] = ' '

        wb = load_workbook(f'{self.save_path}/{self.uniprot_result["proteinName"]}.xlsx')
        result = wb.active

        r = 5
        pep_flag = False
        for i in prange(len(self.peptides)):
            count_peptide = 0
            if self.peptides[i] in self.sequence:
                pep_flag = True
                first_r = r
                not_checked = 0
                for col in prange(len(self.columns_peptides)):
                    if self.columns_peptides[col] == ' ':
                        not_checked += 1
                    else:
                        result.cell(row=r, column=col + 2 - not_checked).value = self.columns_peptides[col]
                        result.cell(row=r, column=col + 2 - not_checked).font = Font(bold=True)
                        result.cell(row=r, column=col + 2 - not_checked).fill = PatternFill(
                            start_color="DAEEF3",
                            end_color="DAEEF3",
                            fill_type="solid")
                        result.cell(row=r, column=col + 2 - not_checked).alignment = Alignment(
                            horizontal='left',
                            vertical='top')
                        result.row_dimensions[r].height = 30
                sdvig = not_checked
                not_checked = 0
                position_col = []
                for col in prange(len(self.columns_aminos)):
                    if self.columns_aminos[col] == ' ':
                        not_checked += 1
                        position_col.append(-1)
                    else:
                        result.cell(row=r, column=col + 8 - sdvig - not_checked).value = self.columns_aminos[col]
                        result.cell(row=r, column=col + 8 - sdvig - not_checked).font = Font(bold=True)
                        result.cell(row=r, column=col + 8 - sdvig - not_checked).fill = PatternFill(
                            start_color="EBF1DE",
                            end_color="EBF1DE",
                            fill_type="solid")
                        result.cell(row=r, column=col + 8 - sdvig - not_checked).alignment = Alignment(
                            horizontal='left',
                            vertical='top')
                        result.row_dimensions[r].height = 30
                        position_col.append(col + 8 - sdvig - not_checked)

                result.cell(row=r, column=1).fill = PatternFill(
                    start_color="DAEEF3",
                    end_color="DAEEF3",
                    fill_type="solid")

                for j in prange(len(self.sequence)):
                    if self.sequence[j: j + len(self.peptides[i])] == self.peptides[i]:
                        count_peptide += 1
                        leftpos = j + 1
                        rigthpos = j + len(self.peptides[i])
                        if j == 0:
                            nter = 'None'
                            cter = self.sequence[j + len(self.peptides[i])]
                        elif j == len(self.sequence) - 1:
                            nter = self.sequence[j - 1]
                            cter = 'None'
                        else:
                            nter = self.sequence[j - 1]
                            cter = self.sequence[j + len(self.peptides[i])]
                        if position_col[0] != -1:
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[0]).value = f'{leftpos}-{rigthpos}'
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[0]).alignment = Alignment(
                                horizontal='left',
                                vertical='top')
                        if position_col[1] != -1:
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[1]).value = nter
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[1]).alignment = Alignment(
                                horizontal='left',
                                vertical='top')
                        if position_col[2] != -1:
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[2]).value = cter
                            result.cell(
                                row=r + count_peptide,
                                column=position_col[2]).alignment = Alignment(
                                horizontal='left',
                                vertical='top')
                        j += len(self.peptides[i])

                category = 1200
                peptide_id = i + 1
                relative = round((count_peptide * 1000) / len(self.sequence), 2)
                peptide = self.peptides[i]
                length_peptide = len(peptide)
                value_list = [category, peptide_id, peptide, length_peptide, count_peptide, relative]

                for i, key in enumerate(self.excelFilters.keys(), start=0):
                    if (i > 9 and i < 16) and not(self.excelFilters[key]):
                        value_list[i - 10] = ' '

                is_checked = 0
                for val in prange(len(value_list)):
                    if value_list[val] == ' ':
                        is_checked += 1
                    else:
                        result.cell(row=first_r + 1, column=val + 2 - is_checked).value = value_list[val]
                        result.cell(row=first_r + 1, column=val + 2 - is_checked).alignment = Alignment(
                            horizontal='left',
                            vertical='top'
                        )

                r += count_peptide + 1

        if not pep_flag:
            result.cell(row=r + 2, column=2).value = 'Not peptides in file'
            result.cell(row=r + 2, column=2).font = Font(bold=True)

        for col in result.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except BaseException:
                    pass

            adjusted_width = (max_length + 2) * 1.2
            result.column_dimensions[column].width = adjusted_width

        result.column_dimensions['A'].width = 4
        wb.save(f'{self.save_path}/{self.uniprot_result["proteinName"]}.xlsx')


class ErrorsHandler:
    pass


class AsyncParser:
    def __init__(self, proteins: List):
        self.proteins = proteins
        self.parsing_result = []
        self.sequences = []

    def run_async_parsing(self):
        try:
            asyncio.run(self.gather_data())
            return (self.parsing_result, self.sequences)
        except BaseException:
            pass

    async def gather_data(self):
       async with aiohttp.ClientSession() as session:
            tasks = []
            for protein in self.proteins:
                task = asyncio.create_task(self.get_json(session, protein))
                tasks.append(task)

            await asyncio.gather(*tasks)

    async def get_json(self, session, protein):
        urluniprot = f'https://rest.uniprot.org/uniprotkb/{protein}.json'
        async with session.get(url=urluniprot) as response:
            response_json = await response.json()
            keys = response_json.keys()

            if "messages" in keys and "url" in keys:
                with open("errorLogs.json", "r") as errorLogs:
                    errors_data = json.load(errorLogs)
                    errors_data["missing"]["proteins"].append(protein)
                with open("errorLogs.json", "w") as errorLogs:
                    json.dump(errors_data, errorLogs, indent=4)
                return False

            result_dict = {
                "proteinName": "None",
                "entryName": "None",
                "entryType": "None",
                "fullName": "None",
                "scientificName": "None",
                "commonName": "None",
                "genes": "None",
                "proteinExistence": "None",
                "Length": "None",
                "massDa": "None",
            }

            protein_json = response_json

            result_dict["proteinName"] = protein_json["primaryAccession"]

            result_dict["entryName"] = protein_json["uniProtkbId"]

            result_dict["fullName"] = protein_json["proteinDescription"]["recommendedName"]["fullName"]["value"]

            result_dict["scientificName"] = protein_json["organism"]["scientificName"]

            result_dict["genes"] = protein_json["genes"][0]["geneName"]["value"]

            temp = str(protein_json["sequence"]["molWeight"])
            result_dict["massDa"] = f'{temp[0:2]},{temp[2:len(temp)]}'

            entryType = protein_json["entryType"]
            result_dict["entryType"] = entryType[entryType.find('(') + 1: len(entryType) - 1]

            result_dict["commonName"] = protein_json["organism"]["commonName"]

            result_dict["proteinExistence"] = protein_json["proteinExistence"][3:len(protein_json["proteinExistence"])]

            sequence = protein_json["sequence"]["value"]

            result_dict["Length"] = protein_json["sequence"]["length"]

            self.parsing_result.append(result_dict)
            self.sequences.append(sequence)
            print(f'Обработал {result_dict["proteinName"]}')


def async_creater(proteinName):
    createrThread = threading.Thread(target=Creater(proteinName).creating_excel())
    createrThread.start()


def async_writer(uniprot_result, sequence):
    writer = Writer(uniprot_result=uniprot_result, sequence=sequence)
    writerThread_main = threading.Thread(target=writer.filling_main_info())
    writerThread_main_peptides = threading.Thread(target=writer.filling_peptides_info())
    writerThread_main.start()
    writerThread_main_peptides.start()


def main():
    with open("config.json", "r") as config:
        config_data = json.load(config)
    parser = AsyncParser(config_data["proteins"]["value"])
    peptides_json, sequences = parser.run_async_parsing()
    for i in range(len(peptides_json)):
        uniprot_result = peptides_json[i]
        sequence = sequences[i]
        async_creater(proteinName=uniprot_result["proteinName"])
        async_writer(uniprot_result=uniprot_result, sequence=sequence)


import time
st = time.time()
main()
print(time.time() - st)

exit()



